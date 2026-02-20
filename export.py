"""
export.py — Excel export for Containment Division Calculator
"""
import io
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")


def _fmt_sheet(ws, col_widths=None):
    """Apply header formatting and auto-width to a worksheet."""
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # Alternate row shading
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill.fill_type == "none":
                    cell.fill = _ALT_FILL


def build_excel(assumptions: dict, headcount_plan: list,
                weekly_df: pd.DataFrame, monthly_df: pd.DataFrame,
                quarterly_df: pd.DataFrame,
                sensitivity_tables: dict | None = None) -> bytes:
    """
    Build an Excel workbook and return as bytes.

    Parameters
    ----------
    assumptions      : dict of model parameters
    headcount_plan   : list[int] of 120 inspector counts
    weekly_df        : output from run_model
    monthly_df       : output from run_model
    quarterly_df     : output from run_model
    sensitivity_tables : dict {sheet_name: pd.DataFrame}
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Assumptions ──────────────────────────────────────────────
        a_rows = []
        for k, v in assumptions.items():
            if isinstance(v, date):
                v = str(v)
            a_rows.append({"Parameter": k, "Value": v})
        pd.DataFrame(a_rows).to_excel(writer, sheet_name="Assumptions", index=False)
        _fmt_sheet(writer.sheets["Assumptions"])

        # ── Headcount Plan ───────────────────────────────────────────
        hc_df = pd.DataFrame({
            "Model Month": range(1, 121),
            "Inspectors":  headcount_plan,
        })
        hc_df.to_excel(writer, sheet_name="Headcount Plan", index=False)
        _fmt_sheet(writer.sheets["Headcount Plan"])

        # ── Weekly ───────────────────────────────────────────────────
        weekly_cols = [
            "week_start", "week_end", "year", "month", "month_idx",
            "inspectors", "team_leads", "n_opscoord", "n_fieldsup", "n_regionalmgr",
            "insp_st_hrs", "insp_ot_hrs", "lead_st_hrs", "lead_ot_hrs",
            "insp_rev_st", "insp_rev_ot", "lead_rev_st", "lead_rev_ot", "revenue_wk",
            "insp_labor_st", "insp_labor_ot", "lead_labor_st", "lead_labor_ot",
            "hourly_labor", "salaried_wk", "overhead_wk", "ebitda_wk",
            "statement_amt", "collections",
            "ar_begin", "ar_end",
            "payroll_cash_out", "interest_paid",
            "cash_begin", "loc_draw", "loc_repay", "cash_end",
            "loc_begin", "loc_end",
            "warn_loc_maxed", "warn_neg_ebitda",
        ]
        existing_w = [c for c in weekly_cols if c in weekly_df.columns]
        weekly_df[existing_w].to_excel(writer, sheet_name="Weekly", index=False)
        _fmt_sheet(writer.sheets["Weekly"])

        # ── Monthly ──────────────────────────────────────────────────
        monthly_cols = [
            "period", "year", "month",
            "inspectors_avg", "team_leads_avg",
            "revenue", "insp_rev_st", "insp_rev_ot",
            "hourly_labor", "salaried_cost", "overhead", "total_labor",
            "ebitda", "ebitda_margin",
            "interest", "ebitda_after_interest", "ebitda_ai_margin",
            "statement_amt", "collections",
            "ar_end", "loc_end", "cash_end", "peak_loc_to_date",
            "loc_draw", "loc_repay",
        ]
        existing_m = [c for c in monthly_cols if c in monthly_df.columns]
        monthly_df[existing_m].to_excel(writer, sheet_name="Monthly", index=False)
        _fmt_sheet(writer.sheets["Monthly"])

        # ── Quarterly ────────────────────────────────────────────────
        quarterly_cols = [
            "yr_q",
            "revenue", "hourly_labor", "salaried_cost", "overhead", "total_labor",
            "ebitda", "ebitda_margin",
            "interest", "ebitda_after_interest", "ebitda_ai_margin",
            "collections", "loc_draw", "loc_repay",
            "ar_end", "loc_end", "cash_end", "peak_loc_to_date",
        ]
        existing_q = [c for c in quarterly_cols if c in quarterly_df.columns]
        quarterly_df[existing_q].to_excel(writer, sheet_name="Quarterly", index=False)
        _fmt_sheet(writer.sheets["Quarterly"])

        # ── Sensitivity tables ───────────────────────────────────────
        if sensitivity_tables:
            for sheet_name, sdf in sensitivity_tables.items():
                safe = sheet_name[:31]
                sdf.to_excel(writer, sheet_name=safe, index=False)
                _fmt_sheet(writer.sheets[safe])

    output.seek(0)
    return output.read()
